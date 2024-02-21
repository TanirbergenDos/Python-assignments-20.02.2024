from openpyxl import Workbook
from string import ascii_lowercase
from random import choice, randint
from datetime import date, datetime


# Create a randomly generated name
name = ''
for _ in range(10):
    name += choice(ascii_lowercase)
name = name.capitalize()

wb = Workbook()
ws = wb.active
ws.title = "TDSheet"
ws.cell(row=1, column=1).value = name
current_date = date.today()
ws.cell(row=1, column=2).value = current_date
ws.cell(row=1, column=3).value = datetime.now().time()

filename = name + str(current_date) + str(randint(100, 999)) + ".xlsx"
path = "мои документы/skcu/"
wb.save(path + filename)



